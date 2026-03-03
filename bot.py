#!/usr/bin/env python3
import os
import re
import time
import logging
from io import BytesIO
from typing import Optional, List

import requests
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

BOT_TOKEN = os.environ.get("BOT_TOKEN", "YOUR_TOKEN_HERE")

def extract_article(url: str) -> Optional[str]:
    match = re.search(r'/catalog/(\d+)/', url)
    if match:
        return match.group(1)
    match = re.search(r'(\d{7,12})', url)
    if match:
        return match.group(1)
    return None

def get_basket(vol: int) -> str:
    if vol < 143:    return "01"
    elif vol < 287:  return "02"
    elif vol < 431:  return "03"
    elif vol < 719:  return "04"
    elif vol < 1007: return "05"
    elif vol < 1061: return "06"
    elif vol < 1115: return "07"
    elif vol < 1169: return "08"
    elif vol < 1313: return "09"
    elif vol < 1601: return "10"
    elif vol < 1655: return "11"
    elif vol < 1919: return "12"
    elif vol < 2045: return "13"
    elif vol < 2189: return "14"
    elif vol < 2405: return "15"
    elif vol < 2621: return "16"
    elif vol < 2837: return "17"
    else:            return "18"

def download_photos(article: str) -> List[BytesIO]:
    article_int = int(article)
    vol = article_int // 100000
    part = article_int // 1000
    basket = get_basket(vol)

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "image/webp,image/apng,image/*,*/*;q=0.8",
        "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Referer": "https://www.wildberries.ru/",
        "Origin": "https://www.wildberries.ru"
    }

    photos = []
    for i in range(1, 31):
        url = f"https://basket-{basket}.wbbasket.ru/vol{vol}/part{part}/{article}/images/big/{i}.webp"
        try:
            r = requests.get(url, headers=headers, timeout=15)
            if r.status_code == 200 and len(r.content) > 1000:
                img_bytes = BytesIO(r.content)
                pil_img = PILImage.open(img_bytes)
                png_bytes = BytesIO()
                pil_img.convert("RGB").save(png_bytes, format="PNG")
                png_bytes.seek(0)
                photos.append(png_bytes)
            else:
                break
        except Exception as e:
            logger.error(f"Photo {i} error: {e}")
            break
        time.sleep(0.1)

    return photos


def create_excel(data: list) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Воронки конкурентов"

    IMG_HEIGHT = 300
    ROW_HEIGHT_PT = IMG_HEIGHT * 0.75
    COL_WIDTH_CHARS = 34

    header_fill = PatternFill("solid", fgColor="1A1A18")
    header_font = Font(name="Calibri", bold=True, color="F5F2EC", size=13)
    accent_font = Font(name="Calibri", bold=True, color="CB11AB", size=16)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="D4CFC7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 50

    for col, title in enumerate(["№", "Ссылка"], 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    max_photos = max((len(e["photos"]) for e in data), default=0)

    for p in range(max_photos):
        col = p + 3
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = COL_WIDTH_CHARS
        cell = ws.cell(row=1, column=col, value=f"Фото {p+1}")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for entry in data:
        row = entry["num"] + 1
        ws.row_dimensions[row].height = ROW_HEIGHT_PT

        c1 = ws.cell(row=row, column=1, value=entry["num"])
        c1.font = accent_font
        c1.alignment = center
        c1.border = border

        url = entry["url"]
        c2 = ws.cell(row=row, column=2, value=url)
        c2.font = Font(name="Calibri", color="CB11AB", size=11, underline="single")
        c2.hyperlink = url
        c2.alignment = left
        c2.border = border

        for p_idx, photo_bytes in enumerate(entry["photos"]):
            col = p_idx + 3
            col_letter = get_column_letter(col)
            cell_addr = f"{col_letter}{row}"
            ws.cell(row=row, column=col).border = border
            try:
                photo_bytes.seek(0)
                xl_img = XLImage(photo_bytes)
                orig_w, orig_h = xl_img.width, xl_img.height
                if orig_h > 0:
                    scale = IMG_HEIGHT / orig_h
                    xl_img.height = IMG_HEIGHT
                    xl_img.width = int(orig_w * scale)
                ws.add_image(xl_img, cell_addr)
            except Exception as e:
                ws.cell(row=row, column=col, value="[ошибка фото]")

        for p_idx in range(len(entry["photos"]), max_photos):
            col = p_idx + 3
            ws.cell(row=row, column=col).border = border

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 Привет! Я парсю фото карточек Wildberries и делаю Excel.\n\n"
        "Просто отправь мне ссылки на товары WB — каждую с новой строки.\n\n"
        "Пример:\n"
        "https://www.wildberries.ru/catalog/123456789/detail.aspx\n"
        "https://www.wildberries.ru/catalog/987654321/detail.aspx"
    )

async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📋 Как пользоваться:\n\n"
        "1. Отправь ссылки на карточки WB (каждую с новой строки)\n"
        "2. Подожди ~30 секунд\n"
        "3. Получи Excel файл с фотографиями всех товаров\n\n"
        "В файле:\n"
        "• Каждый конкурент = одна строка\n"
        "• Все его фото идут по столбцам\n"
        "• Фото крупные, удобно сравнивать"
    )

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    urls = [l for l in lines if 'wildberries.ru' in l or re.search(r'\d{7,12}', l)]

    if not urls:
        await update.message.reply_text(
            "❌ Не нашёл ссылок на WB.\n\n"
            "Отправь ссылки в таком формате:\n"
            "https://www.wildberries.ru/catalog/123456789/detail.aspx"
        )
        return

    msg = await update.message.reply_text(f"⏳ Обрабатываю {len(urls)} товар(ов)...")

    data = []
    for i, url in enumerate(urls, 1):
        await msg.edit_text(f"⏳ Скачиваю фото {i}/{len(urls)}...")
        article = extract_article(url)
        if not article:
            data.append({"num": i, "url": url, "photos": []})
            continue
        photos = download_photos(article)
        logger.info(f"Article {article}: {len(photos)} photos downloaded")
        data.append({"num": i, "url": url, "photos": photos})
        time.sleep(0.3)

    await msg.edit_text("📊 Создаю Excel файл...")

    try:
        excel = create_excel(data)
        total_photos = sum(len(e["photos"]) for e in data)
        await update.message.reply_document(
            document=excel,
            filename="воронки-конкурентов.xlsx",
            caption=f"✅ Готово!\n\nКонкурентов: {len(data)}\nФото всего: {total_photos}"
        )
        await msg.delete()
    except Exception as e:
        await msg.edit_text(f"❌ Ошибка: {e}")


def main():
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_cmd))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    logger.info("Бот запущен!")
    app.run_polling()


if __name__ == "__main__":
    main()
